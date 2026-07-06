from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from pathlib import Path


def save_articles_to_excel(articles, file_path="articles.xlsx"):
    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "通知公告"

    headers = ["标题", "链接", "发布日期", "一句话概括", "要点"]
    sheet.append(headers)

    for article in articles:
        title = article.get("title", "")
        link = article.get("link", "")
        date = article.get("date", "")
        summary = article.get("summary", "")
        key_points = article.get("key_points", [])

        if isinstance(key_points, list):
            key_points_text = "\n".join(key_points)
        else:
            key_points_text = str(key_points)

        row = [title, link, date, summary, key_points_text]
        sheet.append(row)

        link_cell = sheet.cell(row=sheet.max_row, column=2)
        link_cell.hyperlink = link
        link_cell.style = "Hyperlink"

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 45
    sheet.column_dimensions["C"].width = 15
    sheet.column_dimensions["D"].width = 40
    sheet.column_dimensions["E"].width = 50

    workbook.save(file_path)
